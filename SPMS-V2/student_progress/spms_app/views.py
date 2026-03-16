from collections import defaultdict
import datetime
import io
import base64
import matplotlib
matplotlib.use('Agg')   # <---- CRITICAL FIX FOR DJANGO + WINDOWS
import matplotlib.pyplot as plt

import matplotlib.pyplot as plt

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.db.models import IntegerField
from django.db.models.functions import Cast
from django.contrib import messages

from .models import Student, Teacher, Attendance, Subject, Marks

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from django.contrib import messages
from django.db.models.functions import Cast
from django.core.paginator import Paginator

from django.core.paginator import Paginator
import math
import io, base64
import matplotlib.pyplot as plt


import random
from django.core.mail import send_mail
from django.shortcuts import render, redirect
from .models import Teacher, TeacherOTP

from django.contrib import messages
import random
from django.core.mail import send_mail
from django.contrib import messages
from django.shortcuts import render, redirect
from .models import TeacherOTP
from django.utils.timezone import now
from .models import Teacher, TeacherOTP

from django.utils.timezone import now, timedelta
from django.contrib import messages
import random


def home(request):
    return render(request, 'home.html')

def teacher_register(request):
    if request.method == "POST":
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']

        # Check if username already exists
        if Teacher.objects.filter(username=username).exists():
            messages.error(request, "Username already taken!")
            return redirect('teacher_register')

        # Check if email already exists
        if Teacher.objects.filter(email=email).exists():
            messages.error(request, "Email already registered!")
            return redirect('teacher_register')

        # Create teacher directly (NO OTP)
        Teacher.objects.create(
            username=username,
            email=email,
            password=password
        )

        messages.success(request, "Registration successful! You can login now.")
        return redirect('teacher_login')

    return render(request, 'teacher_register.html')

# ---------- TEACHER LOGIN ----------
def teacher_login(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']

        try:
            teacher = Teacher.objects.get(username=username, password=password)
            request.session['teacher_id'] = teacher.id
            return redirect('teacher_dashboard')
        except:
            return render(request, 'teacher_login.html', {'error': 'Invalid login'})

    return render(request, 'teacher_login.html')


# ---------- TEACHER DASHBOARD ----------
def teacher_dashboard(request):
    teacher_id = request.session.get('teacher_id')

    if not teacher_id:
        return redirect('teacher_login')

    teacher = Teacher.objects.get(id=teacher_id)
    return render(request, 'teacher_dashboard.html', {'teacher': teacher})




def manage_students(request):
    teacher_id = request.session.get('teacher_id')

    if not teacher_id:
        return redirect('teacher_login')

    teacher = Teacher.objects.get(id=teacher_id)

    if request.method == "POST":
        name = request.POST.get('name')
        roll_no = request.POST.get('roll_no')

    # --- SAFETY CHECK BEFORE CREATE ---
        if Student.objects.filter(teacher=teacher, roll_no=roll_no).exists():
            messages.error(request, "❌ This roll number already exists!")
            return redirect('manage_students')

    # generate password
        password = f"{roll_no}{teacher.student_default_pass}"

        try:
            Student.objects.create(
                teacher=teacher,
                name=name,
                roll_no=roll_no,
                password=password
            )
            messages.success(request, "✅ Student added successfully!")

        except Exception as e:
            messages.error(request, "❌ Error adding student — duplicate roll detected.")
    
        return redirect('manage_students')
    
    students = Student.objects.filter(teacher=teacher).annotate(
        roll_int=Cast('roll_no', IntegerField())
    ).order_by('roll_int')

    return render(request, 'manage_students.html', {
        'students': students
    })



# ---------- ATTENDANCE DATE LIST ----------
def attendance_list(request):
    teacher_id = request.session.get('teacher_id')

    if not teacher_id:
        return redirect('teacher_login')

    # Get unique dates only
    dates = Attendance.objects.filter(
        teacher_id=teacher_id
    ).values_list('date', flat=True).distinct().order_by('-date')

    return render(request, 'attendance_list.html', {'dates': dates})


# ---------- ADD ATTENDANCE (Select Date) ----------
def add_attendance(request):
    teacher_id = request.session.get('teacher_id')

    if not teacher_id:
        return redirect('teacher_login')

    if request.method == "POST":
        date = request.POST['date']

        # Check if attendance already exists for this date
        exists = Attendance.objects.filter(
            teacher_id=teacher_id,
            date=date
        ).exists()

        if exists:
            messages.error(request, "Attendance for this date is already taken!")
            return redirect('attendance_list')

        request.session['attendance_date'] = date
        return redirect('mark_attendance')

    return render(request, 'add_attendance.html')


# ---------- MARK ATTENDANCE (Checkbox page) ----------
def mark_attendance(request):
    teacher_id = request.session.get('teacher_id')
    date = request.session.get('attendance_date')

    if not teacher_id or not date:
        return redirect('attendance_list')

    teacher = Teacher.objects.get(id=teacher_id)

    # IMPORTANT: sort students by roll number
    students = Student.objects.filter(teacher=teacher).annotate(
    roll_int=Cast('roll_no', IntegerField())
).order_by('roll_int')

    if request.method == "POST":
        present_ids = request.POST.getlist('present')

        for s in students:
            status = "Present" if str(s.id) in present_ids else "Absent"

            Attendance.objects.create(
                teacher=teacher,
                student=s,
                date=date,
                status=status
            )

        return redirect('attendance_list')

    return render(request, 'mark_attendance.html', {
        'students': students,
        'date': date
    })


# ---------- VIEW FULL ATTENDANCE FOR A DATE ----------

def view_attendance_date(request, date):
    teacher_id = request.session.get('teacher_id')

    if not teacher_id:
        return redirect('teacher_login')

    records = Attendance.objects.filter(
        teacher_id=teacher_id,
        date=date
    ).annotate(
        roll_int=Cast('student__roll_no', IntegerField())
    ).order_by('roll_int')

    return render(request, 'view_attendance_date.html', {
        'records': records,
        'date': date
    })


def edit_attendance(request, date):
    teacher_id = request.session.get('teacher_id')

    if not teacher_id:
        return redirect('teacher_login')

    teacher = Teacher.objects.get(id=teacher_id)

    # get students in numeric roll order
    students = Student.objects.filter(teacher=teacher).annotate(
        roll_int=Cast('roll_no', IntegerField())
    ).order_by('roll_int')

    # existing attendance records for this date
    records = Attendance.objects.filter(
        teacher=teacher,
        date=date
    )

    present_ids = set(records.filter(status="Present")
                      .values_list('student_id', flat=True))

    if request.method == "POST":
        selected = request.POST.getlist('present')

        # delete old records for that date
        Attendance.objects.filter(
            teacher=teacher,
            date=date
        ).delete()

        # create fresh updated records
        for s in students:
            status = "Present" if str(s.id) in selected else "Absent"
            Attendance.objects.create(
                teacher=teacher,
                student=s,
                date=date,
                status=status
            )

        return redirect('attendance_list')  # or /attendance/

    return render(request, 'edit_attendance.html', {
        'students': students,
        'date': date,
        'present_ids': present_ids
    })

def delete_attendance(request, date):
    teacher_id = request.session.get('teacher_id')

    if not teacher_id:
        return redirect('teacher_login')

    teacher = Teacher.objects.get(id=teacher_id)

    # delete all attendance records for this date & teacher
    Attendance.objects.filter(
        teacher=teacher,
        date=date
    ).delete()

    messages.success(request, f"Attendance for {date} deleted successfully.")

    return redirect('attendance_list')  # or '/attendance/'


# ---------- SUBJECT LIST (Add / Remove Subjects) ----------
def subject_list(request):
    teacher_id = request.session.get('teacher_id')

    if not teacher_id:
        return redirect('teacher_login')

    teacher = Teacher.objects.get(id=teacher_id)

    if request.method == "POST":
        subject_name = request.POST['subject_name']
        Subject.objects.create(teacher=teacher, subject_name=subject_name)

    subjects = Subject.objects.filter(teacher=teacher)
    return render(request, 'subject_list.html', {'subjects': subjects})


def delete_subject(request, id):
    Subject.objects.get(id=id).delete()
    return redirect('subject_list')


# ---------- ENTER MARKS FOR A SUBJECT ----------
def enter_marks(request, subject_id):
    teacher_id = request.session.get('teacher_id')

    if not teacher_id:
        return redirect('teacher_login')

    teacher = Teacher.objects.get(id=teacher_id)
    subject = get_object_or_404(Subject, id=subject_id)

    students = Student.objects.filter(teacher=teacher).annotate(
    roll_int=Cast('roll_no', IntegerField())
).order_by('roll_int')

    if request.method == "POST":
        for s in students:
            marks_value = request.POST.get(f'marks_{s.id}')

            if marks_value:
                # FIRST: delete old duplicates if any
                Marks.objects.filter(
                    teacher=teacher,
                    student=s,
                    subject=subject.subject_name
                ).delete()

                # THEN create fresh single record
                Marks.objects.create(
                    teacher=teacher,
                    student=s,
                    subject=subject.subject_name,
                    marks=int(marks_value)
                )

        return redirect('enter_marks', subject_id=subject.id)

    # -------- GET EXISTING MARKS TO SHOW ON PAGE --------
    marks_dict = {}
    existing_marks = Marks.objects.filter(
        teacher=teacher,
        subject=subject.subject_name
    )

    for m in existing_marks:
        marks_dict[m.student_id] = m.marks

    return render(request, 'enter_marks.html', {
        'students': students,
        'subject': subject,
        'marks_dict': marks_dict
    })



# ---------- STUDENT LOGIN ----------
def student_login(request):
    if request.method == "POST":
        roll_no = request.POST['roll_no']
        password = request.POST['password']

        try:
            student = Student.objects.get(
                roll_no=roll_no,
                password=password
            )
            request.session['student_id'] = student.id
            return redirect('student_dashboard')

        except Student.DoesNotExist:
            return render(request, 'student_login.html', {
                'error': 'Invalid Roll No or Password'
            })

    return render(request, 'student_login.html')


# ---------- STUDENT DASHBOARD ----------

def student_dashboard(request):
    student_id = request.session.get('student_id')

    if not student_id:
        return redirect('student_login')

    student = Student.objects.get(id=student_id)

    # -------- MARKS GRAPH --------
    marks = Marks.objects.filter(student=student)

    subjects = [m.subject for m in marks]
    scores = [m.marks for m in marks]

    plt.figure()
    plt.bar(subjects, scores)
    plt.xlabel("Subjects")
    plt.ylabel("Marks")
    plt.title("Marks Performance")

    buffer = io.BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    marks_chart = base64.b64encode(buffer.getvalue()).decode()
    buffer.close()
    plt.close()

    # -------- ATTENDANCE GRAPH (WEEK VIEW) --------
    attendance = Attendance.objects.filter(student=student).order_by('date')

    dates = [a.date for a in attendance]
    status_values = [1 if a.status == "Present" else 0 for a in attendance]

    plt.figure(figsize=(7, 3.5))
    plt.plot(dates, status_values, marker="o")

    plt.title("Attendance Trend (Weekly View)", fontsize=12)
    plt.xlabel("Week", fontsize=10)
    plt.ylabel("Status", fontsize=10)
    plt.yticks([0, 1], ["Absent", "Present"])

    ax = plt.gca()

    if dates:
        start_date = dates[0]
        week_points = []
        week_labels = []

        for d in dates:
            week_number = math.floor((d - start_date).days / 7) + 1

            if week_number not in week_labels:
                week_labels.append(week_number)
                week_points.append(d)

        ax.set_xticks(week_points)
        ax.set_xticklabels([f"Week {w}" for w in week_labels])

    plt.grid(True, linestyle="--", alpha=0.5)
    plt.tight_layout()

    buffer = io.BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    attendance_chart = base64.b64encode(buffer.getvalue()).decode()
    buffer.close()
    plt.close()

    # -------- ATTENDANCE % --------
    total_days = attendance.count()
    present_days = attendance.filter(status="Present").count()

    attendance_percent = (
        round((present_days / total_days) * 100, 2)
        if total_days > 0 else 0
    )

    # -------- PAGINATION (10 PER PAGE) --------
    attendance_list = Attendance.objects.filter(student=student)\
                        .order_by('-date')

    paginator = Paginator(attendance_list, 10)
    page_number = request.GET.get('page')
    attendance_page = paginator.get_page(page_number)

    return render(request, 'student_dashboard.html', {
        'student': student,
        'marks': marks,
        'attendance_page': attendance_page,
        'attendance_percent': attendance_percent,
        'marks_chart': marks_chart,
        'attendance_chart': attendance_chart,
    })


def teacher_logout(request):
    request.session.flush()
    return redirect('teacher_login')


def student_logout(request):
    request.session.flush()
    return redirect('student_login')




def edit_subject(request, subject_id):
    teacher_id = request.session.get('teacher_id')

    if not teacher_id:
        return redirect('teacher_login')

    subject = get_object_or_404(Subject, id=subject_id, teacher_id=teacher_id)

    if request.method == "POST":
        new_name = request.POST.get('subject_name')

        if new_name:
            subject.subject_name = new_name
            subject.save()
            return redirect('subject_list')   # or /subjects/

    return render(request, 'edit_subject.html', {'subject': subject})

def edit_student(request, student_id):
    teacher_id = request.session.get('teacher_id')

    if not teacher_id:
        return redirect('teacher_login')

    student = get_object_or_404(Student, id=student_id, teacher_id=teacher_id)
    teacher = Teacher.objects.get(id=teacher_id)

    if request.method == "POST":
        new_name = request.POST.get('name')
        new_roll = request.POST.get('roll_no')

        # --- DUPLICATE ROLL CHECK (IMPORTANT) ---
        exists = Student.objects.filter(
            teacher=teacher,
            roll_no=new_roll
        ).exclude(id=student.id).exists()

        if exists:
            messages.error(request, "This roll number already exists!")
            return redirect(f"/edit-student/{student.id}/")

        # update safely
        student.name = new_name
        student.roll_no = new_roll
        student.password = f"{new_roll}{teacher.student_default_pass}"
        student.save()

        messages.success(request, "Student updated successfully.")
        return redirect('manage_students')

    return render(request, 'edit_student.html', {'student': student})

def delete_student(request, student_id):
    teacher_id = request.session.get('teacher_id')

    if not teacher_id:
        return redirect('teacher_login')

    student = get_object_or_404(Student, id=student_id, teacher_id=teacher_id)
    student.delete()

    messages.success(request, "Student deleted successfully.")
    return redirect('manage_students')



def download_attendance_excel(request):
    teacher_id = request.session.get('teacher_id')

    if not teacher_id:
        return redirect('teacher_login')

    teacher = Teacher.objects.get(id=teacher_id)

    # -------- GET DATA --------
    students = Student.objects.filter(teacher=teacher).annotate(
        roll_int=Cast('roll_no', IntegerField())
    ).order_by('roll_int')

    records = Attendance.objects.filter(teacher=teacher).order_by('date')

    # group dates by month
    month_dates = defaultdict(list)
    all_dates = sorted(set(records.values_list('date', flat=True)))

    for d in all_dates:
        month_dates[d.strftime("%Y-%m")].append(d)

    # lookup: (student_id, date) -> status
    attendance_map = {}
    for r in records:
        attendance_map[(r.student_id, r.date)] = r.status

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    # -------- STYLES --------
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # -------- CREATE SHEET PER MONTH --------
    for month, dates in month_dates.items():

        ws = wb.create_sheet(title=month)

        # -------- HEADER --------
        ws.cell(row=1, column=1, value="Roll No")
        ws.cell(row=1, column=2, value="Student Name")

        col = 3
        for d in sorted(dates):
            ws.cell(row=1, column=col, value=str(d))
            col += 1

        ws.cell(row=1, column=col, value="Attendance %")

        # style header + borders
        for c in range(1, col+1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_all

        # -------- DATA ROWS --------
        row = 2
        for s in students:
            ws.cell(row=row, column=1, value=s.roll_no).border = border_all
            ws.cell(row=row, column=2, value=s.name).border = border_all

            present_count = 0
            total_days = len(dates)

            c = 3
            for d in sorted(dates):
                status = attendance_map.get((s.id, d), "Absent")
                cell = ws.cell(row=row, column=c, value=status)

                if status == "Present":
                    cell.fill = green_fill
                    present_count += 1
                else:
                    cell.fill = red_fill

                cell.alignment = Alignment(horizontal="center")
                cell.border = border_all
                c += 1

            percent = round((present_count / total_days) * 100, 2) if total_days > 0 else 0
            p_cell = ws.cell(row=row, column=c, value=f"{percent}%")
            p_cell.alignment = Alignment(horizontal="center")
            p_cell.border = border_all

            row += 1

        # -------- COLUMN WIDTH --------
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25

        for i in range(3, 3 + len(dates)):
            ws.column_dimensions[chr(64+i)].width = 14

    # -------- RESPONSE --------
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = f"attendance_full_{datetime.date.today()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


def download_marksheet_pdf(request):
    student_id = request.session.get('student_id')

    if not student_id:
        return redirect('student_login')

    student = Student.objects.get(id=student_id)
    teacher = student.teacher

    marks = Marks.objects.filter(student=student)

    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="marksheet_{student.roll_no}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("<b>STUDENT MARKSHEET</b>", styles['Title']))
    elements.append(Spacer(1, 10))

    # Student info
    elements.append(Paragraph(f"<b>Name:</b> {student.name}", styles['Normal']))
    elements.append(Paragraph(f"<b>Roll No:</b> {student.roll_no}", styles['Normal']))
    elements.append(Paragraph(f"<b>Teacher:</b> {teacher.username}", styles['Normal']))
    elements.append(Spacer(1, 15))

    # Table data
    data = [["Subject", "Marks (out of 100)"]]

    total = 0
    count = 0

    for m in marks:
        # ✅ FIX: subject is a string, not FK
        data.append([m.subject, str(m.marks)])
        total += m.marks
        count += 1

    percent = round(total / count, 2) if count > 0 else 0

    data.append(["Total Marks", str(total)])
    data.append(["Percentage", f"{percent}%"])

    # Create table
    table = Table(data, colWidths=[250, 200])

    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightblue),
        ('ALIGN', (1,1), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('PADDING', (0,0), (-1,-1), 8),
    ]))

    elements.append(table)
    doc.build(elements)

    return response

def verify_otp(request):
    if request.method == "POST":
        entered_otp = request.POST['otp']
        temp_data = request.session.get('temp_teacher')

        if not temp_data:
            messages.error(request, "Session expired. Please register again.")
            return redirect('teacher_register')

        email = temp_data['email']

        try:
            otp_record = TeacherOTP.objects.get(email=email)

            if otp_record.is_expired():
                otp_record.delete()
                messages.error(request, "OTP expired. Please register again.")
                return redirect('teacher_register')

            if otp_record.otp == entered_otp:
                Teacher.objects.create(
                    username=temp_data['username'],
                    email=temp_data['email'],
                    password=temp_data['password'],
                    student_default_pass="default123"
                )

                otp_record.delete()
                del request.session['temp_teacher']

                messages.success(request, "Registration successful! Login now.")
                return redirect('teacher_login')

            else:
                messages.error(request, "Invalid OTP. Try again.")

        except TeacherOTP.DoesNotExist:
            messages.error(request, "OTP not found. Please register again.")

    return render(request, 'verify_otp.html')

def resend_otp(request):
    temp_data = request.session.get('temp_teacher')

    if not temp_data:
        return redirect('teacher_register')

    email = temp_data['email']

    try:
        otp_obj = TeacherOTP.objects.get(email=email)
    except TeacherOTP.DoesNotExist:
        messages.error(request, "OTP session expired. Register again.")
        return redirect('teacher_register')

    # Check 60-second limit
    time_diff = now() - otp_obj.created_at

    if time_diff < timedelta(seconds=60):
        messages.error(request, "Wait 60 seconds before resending OTP.")
        return redirect('verify_otp')

    # Generate new OTP
    new_otp = str(random.randint(100000, 999999))

    # Update existing OTP instead of creating new row
    otp_obj.otp = new_otp
    otp_obj.created_at = now()
    otp_obj.save()

    # Send email via Zoho
    send_mail(
        "Your New OTP",
        f"Your new OTP is: {new_otp}",
        "avadhut.gore24@zohomail.in",
        [email],
        fail_silently=False,
    )

    messages.success(request, "New OTP sent to your email.")
    return redirect('verify_otp')
